import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import io

def exportar_excel(df_actividades, df_resultados, graficos, nombre_proyecto):
    print("Iniciando la exportación a Excel")
    filename = f"{nombre_proyecto}_informe_resultados.xlsx"
    
    try:
        wb = Workbook()
        
        #Actividades
        ws_actividades = wb.active
        ws_actividades.title = 'Actividades'
        
        headers_actividades = ['Actividad', 'Planned Value (PV)', 'Earned Value (EV)', 'Actual Cost (AC)', 'Semana']
        ws_actividades.append(headers_actividades)
        
        for r_idx, row in enumerate(df_actividades.values, 2):  # Comenzar en la fila 2 para dejar espacio para los encabezados
            for c_idx, value in enumerate(row, 1):
                ws_actividades.cell(row=r_idx, column=c_idx, value=value)
        
        #Resultados
        ws_resultados = wb.create_sheet(title='Resultados')
        
        headers_resultados = [''] + df_resultados.columns.tolist()
        ws_resultados.append(headers_resultados)
        
        transposed_data = df_resultados.T
        for col_name in transposed_data.columns:
            ws_resultados.append([col_name] + transposed_data[col_name].tolist())
        
        #Gráficos
        ws_graficos = wb.create_sheet(title='Gráficos')
        
        row_offset = 1
        for graph in graficos:
            print(f"Agregando gráfico: {graph['path']}")
            # Añadir título antes de la imagen
            ws_graficos.cell(row=row_offset, column=1, value=graph['title'])
            row_offset += 1 
            
            img = Image(graph['path'])
            ws_graficos.add_image(img, f'A{row_offset}')
            
            row_offset += (img.height // 20) + 5  
        
        # Guardar el archivo
        wb.save(filename)
        print(f"Archivo Excel guardado como: {filename}")
        return filename
    
    except Exception as e:
        print(f"Error al generar el archivo Excel: {e}")
        return None
